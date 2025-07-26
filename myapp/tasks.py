import os
import csv
from celery import shared_task
from openpyxl import load_workbook
from .models import Applicant
from django.db import models
from django.db.models import CharField, DateTimeField
from dateutil.parser import parse


@shared_task
def process_uploaded_file(file_path):
    _, ext = os.path.splitext(file_path.lower())

    if ext == ".csv":
        process_csv(file_path)
    elif ext in [".xlsx", ".xls"]:
        process_excel(file_path)

    os.remove(file_path)


def process_csv(file_path):
    with open(file_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader)
        headers = [
            h.strip().lower().replace(" ", "_").replace("?", "") for h in headers
        ]

        batch = []
        batch_size = 1000

        for row in reader:
            row_data = {
                headers[i]: val for i, val in enumerate(row) if i < len(headers)
            }
            obj = build_applicant_from_dict(row_data)
            batch.append(obj)
            if len(batch) >= batch_size:
                Applicant.objects.bulk_create(batch)
                batch.clear()

        if batch:
            Applicant.objects.bulk_create(batch)


def process_excel(file_path):
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active

    headers = []
    batch = []
    batch_size = 1000

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            headers = [
                str(cell).strip().lower().replace(" ", "_").replace("?", "")
                for cell in row
            ]
            continue

        row_data = {headers[j]: cell for j, cell in enumerate(row) if j < len(headers)}
        obj = build_applicant_from_dict(row_data)
        batch.append(obj)
        if len(batch) >= batch_size:
            Applicant.objects.bulk_create(batch)
            batch.clear()

    if batch:
        Applicant.objects.bulk_create(batch)

    wb.close()


def build_applicant_from_dict(raw_data):
    field_map = {}
    for f in Applicant._meta.get_fields():
        if not (f.concrete and not f.auto_created):
            continue
        verbose = getattr(f, "verbose_name", f.name)
        norm = verbose.strip().lower().replace(" ", "_").replace("?", "")
        if norm in raw_data and f.name not in raw_data:
            field_map[norm] = f.name

    for k, v in field_map.items():
        raw_data[v] = raw_data.pop(k)

    datetime_fields = [
        f.name
        for f in Applicant._meta.get_fields()
        if isinstance(f, models.DateTimeField)
    ]

    for fld in datetime_fields:
        val = raw_data.get(fld)
        if isinstance(val, str):
            val = val.strip()
            if not val:
                raw_data[fld] = None
            else:
                try:

                    raw_data[fld] = parse(val)
                except (ValueError, TypeError):

                    raw_data[fld] = None
        elif isinstance(val, (int, float)):
            raw_data[fld] = None
    model_fields = [
        f.name
        for f in Applicant._meta.get_fields()
        if f.concrete and not f.auto_created
    ]
    data = {}

    for field in model_fields:
        val = raw_data.get(field, None)

        field_obj = Applicant._meta.get_field(field)
        if isinstance(field_obj, CharField):
            max_len = field_obj.max_length
            if isinstance(val, str) and len(val) > max_len:
                val = val[:max_len]

        data[field] = val

    return Applicant(**data)
